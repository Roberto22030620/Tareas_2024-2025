#Presentar este código en el examen de HPC
import openpyxl, random, threading

workbook = openpyxl.load_workbook("Examen_HPC.xlsx")
hoja = workbook.active

def registrar_venta():
    global hoja

    montos = [random.randint(1000, 5000) for _ in range(3)]
    suma_montos = sum(montos)

    vendedores = random.sample(range(1, 7), 3)

    with threading.Lock():
        fila = hoja.max_row + 1

        hoja.cell(row=fila, column=1, value=suma_montos)

        for col in range(2, 8):
            if col - 1 in vendedores:
                hoja.cell(row=fila, column=col, value="X")
            else:
                hoja.cell(row=fila, column=col, value="")

for _ in range(100):
    hilo = threading.Thread(target=registrar_venta)
    hilo.start()

with threading.Lock():
    workbook.save("Examen_HPC.xlsx")

print("Ventas registradas con éxito.")
